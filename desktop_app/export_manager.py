"""
PharmaPOS NG - Data Export Manager

Handles exporting data to Excel, PDF, and CSV formats.
"""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Dict, Any, Optional
import csv

from desktop_app.logger import get_logger

logger = get_logger(__name__)


class ExportManager:
    """Manages data export to various formats."""
    
    def __init__(self, export_dir: Optional[Path] = None):
        """Initialize export manager.
        
        Args:
            export_dir: Directory for exports (defaults to ./exports)
        """
        from desktop_app.config import PROJECT_ROOT
        
        self.export_dir = export_dir or PROJECT_ROOT / "exports"
        self.export_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"ExportManager initialized: {self.export_dir}")
    
    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        headers: Optional[List[str]] = None
    ) -> tuple[bool, str, Optional[Path]]:
        """Export data to CSV file.
        
        Args:
            data: List of dictionaries to export
            filename: Output filename (without extension)
            headers: Optional list of column headers
            
        Returns:
            Tuple of (success, message, file_path)
        """
        try:
            if not data:
                return False, "No data to export", None
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            csv_filename = f"{filename}_{timestamp}.csv"
            file_path = self.export_dir / csv_filename
            
            # Get headers from first row if not provided
            if headers is None:
                headers = list(data[0].keys())
            
            # Write CSV
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                
                for row in data:
                    # Convert Decimal to float for CSV
                    clean_row = {}
                    for key, value in row.items():
                        if isinstance(value, Decimal):
                            clean_row[key] = float(value)
                        elif isinstance(value, datetime):
                            clean_row[key] = value.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            clean_row[key] = value
                    writer.writerow(clean_row)
            
            logger.info(f"CSV export successful: {file_path}")
            return True, f"Exported to {csv_filename}", file_path
            
        except Exception as e:
            logger.error(f"CSV export failed: {e}", exc_info=True)
            return False, f"Export failed: {str(e)}", None
    
    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        sheet_name: str = "Data",
        headers: Optional[List[str]] = None
    ) -> tuple[bool, str, Optional[Path]]:
        """Export data to Excel file.
        
        Args:
            data: List of dictionaries to export
            filename: Output filename (without extension)
            sheet_name: Excel sheet name
            headers: Optional list of column headers
            
        Returns:
            Tuple of (success, message, file_path)
        """
        try:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                return False, "openpyxl not installed. Run: pip install openpyxl", None
            
            if not data:
                return False, "No data to export", None
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"{filename}_{timestamp}.xlsx"
            file_path = self.export_dir / excel_filename
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Get headers
            if headers is None:
                headers = list(data[0].keys())
            
            # Write headers with styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    
                    # Convert types
                    if isinstance(value, Decimal):
                        value = float(value)
                    elif isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(file_path)
            
            logger.info(f"Excel export successful: {file_path}")
            return True, f"Exported to {excel_filename}", file_path
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}", exc_info=True)
            return False, f"Export failed: {str(e)}", None
    
    def export_to_pdf(
        self,
        title: str,
        data: List[Dict[str, Any]],
        filename: str,
        headers: Optional[List[str]] = None,
        orientation: str = "portrait"
    ) -> tuple[bool, str, Optional[Path]]:
        """Export data to PDF file.
        
        Args:
            title: Report title
            data: List of dictionaries to export
            filename: Output filename (without extension)
            headers: Optional list of column headers
            orientation: Page orientation ('portrait' or 'landscape')
            
        Returns:
            Tuple of (success, message, file_path)
        """
        try:
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import letter, A4, landscape
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            except ImportError:
                return False, "reportlab not installed. Run: pip install reportlab", None
            
            if not data:
                return False, "No data to export", None
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            pdf_filename = f"{filename}_{timestamp}.pdf"
            file_path = self.export_dir / pdf_filename
            
            # Setup document
            pagesize = landscape(letter) if orientation == "landscape" else letter
            doc = SimpleDocTemplate(str(file_path), pagesize=pagesize)
            
            # Container for elements
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Add title
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 0.2 * inch))
            
            # Add generation info
            info_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            elements.append(Paragraph(info_text, styles['Normal']))
            elements.append(Spacer(1, 0.3 * inch))
            
            # Prepare table data
            if headers is None:
                headers = list(data[0].keys())
            
            table_data = [headers]
            
            for row in data:
                row_data = []
                for header in headers:
                    value = row.get(header, "")
                    
                    # Convert types
                    if isinstance(value, Decimal):
                        value = f"{float(value):.2f}"
                    elif isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M")
                    else:
                        value = str(value)
                    
                    row_data.append(value)
                
                table_data.append(row_data)
            
            # Create table
            table = Table(table_data)
            
            # Style table
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            
            logger.info(f"PDF export successful: {file_path}")
            return True, f"Exported to {pdf_filename}", file_path
            
        except Exception as e:
            logger.error(f"PDF export failed: {e}", exc_info=True)
            return False, f"Export failed: {str(e)}", None


__all__ = ['ExportManager']
